# -*- coding: utf-8 -*-
"""
叶子节点扩展脚本
功能：读取Excel中的解法目录树，找到叶子节点，调用千问大模型获取考法，
      然后将考法作为新的子节点插入到Excel中。

依赖安装：pip install openpyxl requests
"""

import openpyxl
import requests
import json
import time
import re

# ====================== 配置区 ======================
EXCEL_PATH = r'c:\Users\win\Desktop\4\result.xlsx'
OUTPUT_PATH = r'c:\Users\win\Desktop\4\result_expanded.xlsx'

# 千问API配置（通义千问 OpenAI兼容模式）
API_KEY = 'sk-44bcd6c4e5c64f7a91ec03789d0eeb27'  # ← 替换为你的API Key
MODEL = 'qwen3.7-plus'
BASE_URL = 'https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions'

# 过滤配置（设为 None 则处理所有叶子节点）
# 例如只处理第十一章：FILTER_CHAPTER = '第十一章 概率与统计'
FILTER_CHAPTER = None  # 设为章节名可只处理该章，None处理全部

# API调用间隔（秒），避免频率限制
API_DELAY = 1
# ====================================================


def build_tree(ws):
    """从Excel构建树结构，处理重名节点（如多个章节都有'数学思想方法'）"""
    rows_data = []
    for r in range(2, ws.max_row + 1):  # 跳过表头
        subject = ws.cell(row=r, column=1).value
        parent_name = ws.cell(row=r, column=2).value
        name = ws.cell(row=r, column=3).value
        rows_data.append({
            'row': r,
            'subject': subject,
            'parent_name': parent_name,
            'name': name,
            'children': [],
            'parent_node': None
        })

    # 按行顺序处理，用 name_to_latest 记录最近出现的同名节点
    # 这样子节点可以正确关联到最近的同名父节点
    name_to_latest = {}
    for node in rows_data:
        if node['parent_name'] is not None and node['parent_name'] in name_to_latest:
            parent = name_to_latest[node['parent_name']]
            parent['children'].append(node)
            node['parent_node'] = parent
        name_to_latest[node['name']] = node

    return rows_data


def get_full_path(node):
    """从叶子节点向上追溯，获取完整路径"""
    path = [node['name']]
    current = node
    while current['parent_node'] is not None:
        current = current['parent_node']
        path.append(current['name'])
    path.reverse()
    return path


def get_chapter(node):
    """获取节点所属的章节（根节点）"""
    current = node
    while current['parent_node'] is not None:
        current = current['parent_node']
    return current['name']


def call_qianwen(prompt, api_key, model):
    """调用千问API"""
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {api_key}'
    }
    data = {
        'model': model,
        'messages': [
            {
                'role': 'system',
                'content': (
                    '你是一个高中数学教育专家，精通高考数学的各种题型和解法。'
                    '请根据给定的知识点路径，列出该知识点在高考中常见的几种解法或考法。'
                    '要求：\n'
                    '1. 只列出解法/考法的名称，每行一个\n'
                    '2. 用数字编号，如 1. xxx  2. xxx\n'
                    '3. 名称要简洁精准，适合作为目录条目\n'
                    '4. 不需要详细解释，只要名称\n'
                    '5. 一般列出3-8个常见考法即可'
                )
            },
            {
                'role': 'user',
                'content': prompt
            }
        ],
        'temperature': 0.7,
        'enable_thinking': False
    }

    response = requests.post(BASE_URL, headers=headers, json=data, timeout=60)
    if response.status_code != 200:
        print(f"  ⚠️ 响应状态码: {response.status_code}")
        print(f"  ⚠️ 响应内容: {response.text[:500]}")
    response.raise_for_status()
    result = response.json()
    content = result['choices'][0]['message']['content']
    # 如果开启了思考模式，thinking内容在单独字段，这里只取最终回答
    return content


def parse_methods(response_text):
    """从API返回的文本中解析出考法列表"""
    methods = []
    lines = response_text.strip().split('\n')
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # 去掉编号：1. / 1、/ 1) / 1） / - / · / • / * 等前缀
        cleaned = re.sub(r'^[\d]+[.、\)）]\s*', '', line)
        cleaned = re.sub(r'^[-·•\*]\s*', '', cleaned)
        cleaned = cleaned.strip()
        # 过滤掉空行和过短的内容（可能是标题或无用内容）
        if cleaned and len(cleaned) >= 2:
            methods.append(cleaned)
    return methods


def main():
    print("=" * 60)
    print("叶子节点扩展工具 - 调用千问大模型")
    print("=" * 60)

    # 1. 读取Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    print(f"\n📖 读取Excel完成，共 {ws.max_row - 1} 行数据")

    # 2. 构建树
    rows_data = build_tree(ws)

    # 3. 找出所有叶子节点
    all_parent_names = set()
    for node in rows_data:
        if node['parent_name'] is not None:
            all_parent_names.add(node['parent_name'])

    leaf_nodes = [node for node in rows_data if node['name'] not in all_parent_names]

    # 4. 按章节过滤（如果设置了过滤条件）
    if FILTER_CHAPTER:
        leaf_nodes = [leaf for leaf in leaf_nodes if get_chapter(leaf) == FILTER_CHAPTER]
        print(f"🔍 过滤章节「{FILTER_CHAPTER}」，找到 {len(leaf_nodes)} 个叶子节点")
    else:
        print(f"🔍 找到 {len(leaf_nodes)} 个叶子节点（全部章节）")

    if not leaf_nodes:
        print("⚠️ 没有找到需要处理的叶子节点！")
        return

    # 预览要处理的叶子节点
    print("\n📋 将要处理的叶子节点：")
    for i, leaf in enumerate(leaf_nodes[:10]):
        path = get_full_path(leaf)
        print(f"  {i+1}. {' > '.join(path)}")
    if len(leaf_nodes) > 10:
        print(f"  ... 还有 {len(leaf_nodes) - 10} 个")

    # 确认
    confirm = input(f"\n确认处理 {len(leaf_nodes)} 个叶子节点？(y/n): ").strip().lower()
    if confirm != 'y':
        print("已取消。")
        return

    # 5. 从后往前处理（避免插入行导致行号偏移）
    leaf_nodes_sorted = sorted(leaf_nodes, key=lambda x: x['row'], reverse=True)

    total = len(leaf_nodes_sorted)
    success_count = 0
    error_count = 0

    for idx, leaf in enumerate(leaf_nodes_sorted):
        full_path = get_full_path(leaf)
        path_str = ' '.join(full_path)

        prompt = f"{path_str} 有哪些几种解法或考法？"

        print(f"\n[{idx+1}/{total}] 🔄 处理: {' > '.join(full_path)}")
        print(f"  提示词: {prompt}")

        try:
            # 调用API
            response_text = call_qianwen(prompt, API_KEY, MODEL)
            print(f"  📝 API返回:\n{''.join(['    ' + l + chr(10) for l in response_text.split(chr(10))[:5]])}")

            # 解析结果
            methods = parse_methods(response_text)
            print(f"  ✅ 解析出 {len(methods)} 个考法: {methods}")

            if methods:
                # 在叶子节点的下一行开始插入
                insert_row = leaf['row'] + 1
                for i, method in enumerate(methods):
                    ws.insert_rows(insert_row + i)
                    ws.cell(row=insert_row + i, column=1, value=leaf['subject'])
                    ws.cell(row=insert_row + i, column=2, value=leaf['name'])  # 父节点=当前叶子
                    ws.cell(row=insert_row + i, column=3, value=method)       # 新叶子=考法

                # 更新后续节点的行号（因为插入了新行）
                num_inserted = len(methods)
                for node in rows_data:
                    if node['row'] >= insert_row:
                        node['row'] += num_inserted

                success_count += 1
            else:
                print(f"  ⚠️ 未解析出有效考法")

            # 避免API频率限制
            time.sleep(API_DELAY)

        except requests.exceptions.HTTPError as e:
            print(f"  ❌ API请求错误: {e}")
            error_count += 1
            # 如果是429（频率限制），等待更久
            if e.response and e.response.status_code == 429:
                print("  ⏳ 触发频率限制，等待10秒...")
                time.sleep(10)
            continue
        except Exception as e:
            print(f"  ❌ 处理出错: {e}")
            error_count += 1
            continue

    # 6. 保存结果
    wb.save(OUTPUT_PATH)
    print(f"\n{'=' * 60}")
    print(f"✅ 完成！")
    print(f"  成功处理: {success_count} 个叶子节点")
    print(f"  处理失败: {error_count} 个")
    print(f"  结果保存到: {OUTPUT_PATH}")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
