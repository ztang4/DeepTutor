# -*- coding: utf-8 -*-
"""
组卷网 解题方法目录爬虫
爬取 https://zujuan.xkw.com/gzsx/jtff181463/o2 的解法目录及所有子目录

使用方法:
    pip install DrissionPage openpyxl
    python scraper.py
"""

from DrissionPage import ChromiumPage, ChromiumOptions
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import json
import time

START_URL = "https://zujuan.xkw.com/gzsx/jtff181463/o2"
SUBJECT = "高中数学"
DELAY = 1.5


def create_browser():
    """创建浏览器"""
    co = ChromiumOptions()
    co.set_argument("--window-size", "1920,1080")
    co.set_argument("--lang=zh-CN")
    return ChromiumPage(co)


def wait_for_page(page, timeout=15):
    """等待页面加载完成"""
    for _ in range(timeout):
        title = page.title
        if title and "请稍候" not in title and len(title) > 2:
            return True
        time.sleep(1)
    return False


def get_node_depth(node):
    """通过遍历DOM父节点，计算 li[tree-id] 的嵌套深度"""
    depth = 0
    try:
        current = node.parent()
        while current:
            tag = current.tag
            if tag and tag.lower() == 'li' and current.attr('tree-id'):
                depth += 1
            cls = current.attr('class') or ''
            if 'tree-box' in cls:
                break
            current = current.parent()
    except Exception:
        pass
    return depth


def click_and_expand_tree(page):
    """尝试展开所有折叠的树节点"""
    expanded = False
    for selector in [
        "css:.tree-box .arrow",
        "css:.tree-box .toggle",
        "css:.tree-box .expand-icon",
        "css:.tree-box i.icon",
    ]:
        try:
            toggles = page.eles(selector)
            if toggles:
                print(f"  🔽 找到 {len(toggles)} 个可展开节点")
                for t in toggles:
                    try:
                        t.click()
                        time.sleep(0.3)
                    except Exception:
                        pass
                expanded = True
        except Exception:
            pass

    try:
        parents = page.eles("css:.tree-box li.has-child, css:.tree-box li.parent")
        if parents:
            for p in parents:
                try:
                    p.click()
                    time.sleep(0.3)
                except Exception:
                    pass
            expanded = True
    except Exception:
        pass

    if expanded:
        time.sleep(1)
    return expanded


def extract_tree(page):
    """提取左侧树形目录，正确计算每个节点的深度"""
    results = []
    try:
        nodes = page.eles("css:.tree-box li[tree-id]")
        if not nodes:
            print("  ⚠ 未找到树形目录节点")
            return results

        for node in nodes:
            tree_id = node.attr("tree-id") or ""
            cls = node.attr("class") or ""

            # 获取节点名称
            link_el = None
            try:
                link_el = node.ele("css:a", timeout=0.5)
            except Exception:
                pass

            if link_el:
                name = link_el.text.strip()
            else:
                name = node.text.strip().split("\n")[0]

            # 计算真实深度
            depth = get_node_depth(node)

            # 是否是父节点
            is_parent = "tree-children" in cls

            if name:
                results.append({
                    "tree_id": tree_id,
                    "name": name,
                    "depth": depth,
                    "is_parent": is_parent,
                })

        print(f"  ✅ 找到 {len(results)} 个节点")
        # 打印树形预览
        for item in results[:20]:
            indent = "    " * item["depth"]
            marker = "📁" if item["is_parent"] else "📄"
            print(f"    {indent}{marker} {item['name']}")
        if len(results) > 20:
            print(f"    ... 还有 {len(results) - 20} 个节点")

    except Exception as e:
        print(f"  ❌ 提取异常: {e}")

    return results


def build_parent_map(tree_items):
    """
    根据深度信息构建父子关系。
    返回列表: [(name, parent_name, depth), ...]
    """
    rows = []
    # 用栈追踪每个深度级别的最近父节点
    parent_stack = []  # [(depth, name), ...]

    for item in tree_items:
        depth = item["depth"]
        name = item["name"]

        # 弹出栈中深度 >= 当前深度的项（它们不是当前节点的父节点）
        while parent_stack and parent_stack[-1][0] >= depth:
            parent_stack.pop()

        parent_name = parent_stack[-1][1] if parent_stack else ""

        rows.append((name, parent_name, depth))

        # 如果是父节点，压入栈
        if item["is_parent"]:
            parent_stack.append((depth, name))

    return rows


def save_excel(tree_items, crawl_info):
    """
    保存为 Excel，3列格式：学科 | 上级目录 | 目录名
    先列所有顶级目录，然后按深度优先展开子目录
    """
    wb = Workbook()

    # ---- 样式 ----
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    normal_font = Font(name="微软雅黑", size=10)
    bold_font = Font(name="微软雅黑", size=10, bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    # ---- Sheet1: 目录树 ----
    ws = wb.active
    ws.title = "解法目录树"

    headers = ["学科", "上级目录", "目录名"]
    col_widths = [12, 30, 40]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w
    ws.freeze_panes = "A2"

    # 构建父子关系
    rows_data = build_parent_map(tree_items)

    # 写入数据：先列所有顶级，再按顺序展开
    # 分离顶级和非顶级
    top_level = [(name, parent, depth) for name, parent, depth in rows_data if depth == 0]
    children = [(name, parent, depth) for name, parent, depth in rows_data if depth > 0]

    # 按原始顺序分组：找到每个顶级节点后面跟着的子节点
    groups = []  # [(top_name, [children...])]
    current_top = None
    current_children = []

    for item in tree_items:
        if item["depth"] == 0:
            if current_top is not None:
                groups.append((current_top, current_children))
            current_top = item["name"]
            current_children = []
        else:
            current_children.append(item)
    if current_top is not None:
        groups.append((current_top, current_children))

    row = 2

    # 先写所有顶级目录
    for top_name, _ in groups:
        ws.cell(row=row, column=1, value=SUBJECT).font = normal_font
        ws.cell(row=row, column=3, value=top_name).font = bold_font
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = thin_border
        if row % 2 == 0:
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = alt_fill
        row += 1

    # 再按顺序展开每个顶级的子目录
    for top_name, group_children in groups:
        if not group_children:
            continue

        # 用栈重建子节点的父关系
        parent_stack = [(0, top_name)]  # 顶级作为初始父节点 (depth=0)

        for child in group_children:
            depth = child["depth"]
            name = child["name"]

            # 弹出栈中深度 >= 当前深度的
            while parent_stack and parent_stack[-1][0] >= depth:
                parent_stack.pop()

            parent_name = parent_stack[-1][1] if parent_stack else top_name

            ws.cell(row=row, column=1, value=SUBJECT).font = normal_font
            ws.cell(row=row, column=2, value=parent_name).font = normal_font
            ws.cell(row=row, column=3, value=name).font = normal_font
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = thin_border
            if row % 2 == 0:
                for c in range(1, 4):
                    ws.cell(row=row, column=c).fill = alt_fill
            row += 1

            if child["is_parent"]:
                parent_stack.append((depth, name))

    # ---- Sheet2: 爬取信息 ----
    ws2 = wb.create_sheet("爬取信息")
    info = [
        ["项目", "内容"],
        ["目标URL", crawl_info.get("start_url", "")],
        ["页面标题", crawl_info.get("page_title", "")],
        ["爬取时间", crawl_info.get("crawl_time", "")],
        ["总节点数", str(len(tree_items))],
    ]
    for col, h in enumerate(info[0], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 55
    for r, (k, v) in enumerate(info[1:], 2):
        ws2.cell(row=r, column=1, value=k).font = bold_font
        ws2.cell(row=r, column=2, value=str(v)).font = normal_font

    excel_path = "result.xlsx"
    wb.save(excel_path)
    print(f"  ✅ Excel 已保存: {excel_path}")


def main():
    print("=" * 55)
    print("  🚀 组卷网解题方法目录爬虫")
    print(f"  📌 目标: {START_URL}")
    print("=" * 55)

    page = create_browser()

    try:
        # 1. 访问页面
        print("\n[1/4] ⏳ 访问页面...")
        page.get(START_URL)
        time.sleep(3)

        if not wait_for_page(page):
            print("  ⚠ 页面可能未完全加载，继续尝试...")

        print(f"  📄 标题: {page.title}")

        # 2. 展开所有节点
        print("\n[2/4] 🔽 展开所有目录节点...")
        click_and_expand_tree(page)
        time.sleep(DELAY)

        # 3. 提取目录树
        print("\n[3/4] 📂 提取解法目录...")
        tree_items = extract_tree(page)

        if not tree_items:
            print("  ❌ 未提取到任何节点，保存页面用于调试...")
            page.get_screenshot(path="screenshot.png", full_page=True)
            with open("page.html", "w", encoding="utf-8") as f:
                f.write(page.html)
            print("  已保存 screenshot.png 和 page.html")
            return

        # 4. 保存结果
        print("\n[4/4] 💾 保存结果...")
        crawl_info = {
            "start_url": START_URL,
            "crawl_time": time.strftime("%Y-%m-%d %H:%M:%S"),
            "page_title": page.title,
        }

        # 保存 JSON
        output = {**crawl_info, "tree": tree_items}
        with open("result.json", "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        print(f"  ✅ result.json")

        # 保存 Excel
        save_excel(tree_items, crawl_info)

        print("\n" + "=" * 55)
        print("  🎉 爬取完成！")
        print(f"  📊 共 {len(tree_items)} 个目录节点")
        print(f"  📗 result.xlsx  - Excel目录树")
        print(f"  📁 result.json  - JSON数据")
        print("=" * 55)

    except Exception as e:
        print(f"\n❌ 出错: {e}")
        import traceback
        traceback.print_exc()
    finally:
        page.quit()
        print("\n👋 完成")


if __name__ == "__main__":
    main()
